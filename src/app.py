import tkinter as tk
from openpyxl import load_workbook
from typing import List

import functions as f

#wb_dir = '../excels/test.xlsx'
#wb = load_workbook(wb_dir)
#ws = wb.active

#assert ws is not None

headers = ["head1", "head2"]

#def init_excel():
    #assert ws is not None
    #for i, h in enumerate(headers, start=1):
        #ws.cell(row=1, column=i, value = h)
    #wb.save(wb_dir)

root = tk.Tk()
root.title("Exo Software")
root.geometry("500x300")
root.config(bg="light blue")

entries: List[tk.Entry] = []

for i, text in enumerate(headers):
    tk.Label(root, text=text).grid(row=i, column=0, padx=10, pady=5, sticky="w")
    entry = tk.Entry(root)
    entry.grid(row=i, column=1, padx=10, pady=5)
    entries.append(entry)


submit_btn = tk.Button(
    root,
    text="Submit",
    #command=lambda: f.insert(wb_dir, entries)
)
submit_btn.grid(row=len(headers), column=1, pady=20)

root.mainloop()
