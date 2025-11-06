from openpyxl import load_workbook

def insert(filepath: str, entries):
    """Insert a row to Excel from a list of Tkinter Entry widgets."""
    wb = load_workbook(filepath)
    ws = wb.active

    assert ws is not None

    next_row = ws.max_row + 1

    for col, entry in enumerate(entries, start=1):
        ws.cell(row=next_row, column=col, value=entry.get())

    wb.save(filepath)

    for entry in entries:
        entry.delete(0, "end")